# -*- coding: utf-8 -*-
"""
[EXPORT] Générateur de Fiche Achat Excel (.xlsx) conforme au modèle FOR-ACH-03-12
=============================================================================
Reconstruit la trame officielle TB Groupe avec openpyxl :
- En-têtes noirs et caractères gras imprimables
- Structure multi-références (Item / Inner / Master / Métal / Manche / Packaging / EAN)
- Bloc réglementaire AGEC intégral (MOAH / MOSH)
"""
from datetime import date
import io
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Bornes de largeur de colonne, en caracteres.
LARGEUR_COLONNE_MIN = 14
LARGEUR_COLONNE_MAX = 45


def generate_fiche_excel_bytes(data: dict[str, Any], items: list[dict[str, Any]]) -> bytes:
    """
    Génère le binaire d'une Fiche Achat Excel (.xlsx) à partir des données formulaires.

    Args:
        data: dictionnaire des champs globaux (supplier, po_number, transport, sample, etc.)
        items: liste des dictionnaires de références (reference, name, dimensions, EAN, etc.)

    Returns:
        Octets (bytes) du fichier .xlsx généré.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Sheet"
    ws.views.sheetView[0].showGridLines = True

    # Styles réutilisables
    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_subhead = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    font_title = Font(name="Arial", size=14, bold=True)
    font_sec_head = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=9, bold=True)
    font_normal = Font(name="Arial", size=9)
    font_small = Font(name="Arial", size=8, italic=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    border_thin_side = Side(style="thin", color="B0B0B0")
    border_all = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)

    current_row = 1

    def write_section_header(title: str) -> None:
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.fill = fill_black
        cell.font = font_sec_head
        cell.alignment = align_left
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    def write_sub_header(cols: list[str]) -> None:
        nonlocal current_row
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = fill_subhead
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_all
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # 1. EN-TÊTE OFFICIEL
    ws.merge_cells("A1:B3")
    logo_cell = ws.cell(row=1, column=1, value="TB GROUPE\nTARRERIAS-BONJEAN")
    logo_cell.font = Font(name="Arial", size=12, bold=True)
    logo_cell.alignment = align_center
    logo_cell.border = border_all

    ws.merge_cells("C1:E3")
    title_cell = ws.cell(row=1, column=3, value="PURCHASE SHEET\nFOR-ACH-03-12")
    title_cell.font = font_title
    title_cell.alignment = align_center
    title_cell.border = border_all

    supplier = data.get("supplier") or ""
    po = data.get("po_number") or ""
    n_lot = data.get("n_lot") or "0000000000"
    today_str = date.today().strftime("%d/%m/%Y")

    ws.merge_cells("F1:G3")
    info_text = f"N° LOT : {n_lot}\nDate d'application : 15/05/2023\nDate de commande : {today_str}\nN° PO : {po}"
    info_cell = ws.cell(row=1, column=6, value=info_text)
    info_cell.font = font_normal
    info_cell.alignment = align_left
    info_cell.border = border_all

    current_row = 5

    # 2. DESCRIPTION OF THE PRODUCT / SUPPLIER
    write_section_header("DESCRIPTION OF THE PRODUCT / SUPPLIER")
    ws.cell(row=current_row, column=1, value="SUPPLIER:").font = font_bold
    ws.cell(row=current_row, column=2, value=supplier).font = font_normal
    current_row += 1

    items_to_render = items if items else [{"reference": data.get("code_article", ""), "name": data.get("designation", "")}]

    write_sub_header(["REFERENCE", "NAME (EN)", "DESIGNATION FR", "", "", "", ""])
    for it in items_to_render:
        ws.cell(row=current_row, column=1, value=str(it.get("reference", ""))).alignment = align_center
        ws.cell(row=current_row, column=2, value=str(it.get("name", ""))).alignment = align_left
        ws.cell(row=current_row, column=3, value=str(it.get("french_desc", ""))).alignment = align_left
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_all
        current_row += 1
    current_row += 1

    # 3. TRANSPORT
    write_section_header("TRANSPORT")
    trans_info = (
        f"FORWARDER: {data.get('forwarder', 'QUALITAIRSEA')}\n"
        f"TYPE OF TRANSPORT: {data.get('transport_type', 'Sea')}\n"
        f"PORT / DESTINATION: {data.get('port', 'FOS SUR MER')}\n"
        f"ETD: {data.get('etd', '')}"
    )
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=7)
    tc = ws.cell(row=current_row, column=1, value=trans_info)
    tc.font = font_normal
    tc.alignment = align_left
    current_row += 4

    # 4. PACKAGING INFORMATIONS
    write_section_header("PACKAGING INFORMATIONS")
    write_sub_header(["REFERENCE", "ITEM (pcs)", "INNER (pcs)", "MASTER (pcs)", "PACKAGING", "", ""])
    for it in items_to_render:
        ws.cell(row=current_row, column=1, value=str(it.get("reference", ""))).alignment = align_center
        ws.cell(row=current_row, column=2, value=str(it.get("pcs_item", 1))).alignment = align_center
        ws.cell(row=current_row, column=3, value=str(it.get("pcs_inner", ""))).alignment = align_center
        ws.cell(row=current_row, column=4, value=str(it.get("pcs_master", ""))).alignment = align_center
        ws.cell(row=current_row, column=5, value=str(it.get("packaging_desc", ""))).alignment = align_left
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_all
        current_row += 1

    pkg_details = (
        f"SHIPPING MARKS: {data.get('shipping_marks', '')}\n"
        f"PARTICULAR DOC REQUIRED: {data.get('particular_doc', '')}"
    )
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=7)
    ws.cell(row=current_row, column=1, value=pkg_details).font = font_normal
    current_row += 3

    # 5. SAMPLE
    write_section_header("SAMPLE")
    send_us = ", ".join(data.get("please_send_us", [])) if isinstance(data.get("please_send_us"), list) else str(data.get("please_send_us") or "")
    sample_text = (
        f"TESTING SAMPLES (BEFORE LAUNCHING PRODUCTION): {'Yes' if data.get('testing_samples') else 'No'}\n"
        f"SHIPPING PAID BY: {data.get('shipping_paid_by', 'TB')}\n"
        f"PLEASE SEND US: {send_us}"
    )
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=7)
    ws.cell(row=current_row, column=1, value=sample_text).font = font_normal
    current_row += 3

    # 6. PRODUCT MATERIAL (METAL & HANDLE)
    write_section_header("PRODUCT MATERIAL")
    ws.cell(row=current_row, column=1, value="METAL PART INFORMATIONS").font = font_bold
    current_row += 1
    write_sub_header(["REFERENCE", "THICKNESS (mm)", "LENGTH (mm)", "QUALITY", "CHROME %", "HEAT TREATMENT", "FINISHING"])
    for it in items_to_render:
        ws.cell(row=current_row, column=1, value=str(it.get("reference", ""))).alignment = align_center
        ws.cell(row=current_row, column=2, value=str(it.get("thickness", ""))).alignment = align_center
        ws.cell(row=current_row, column=3, value=str(it.get("length", ""))).alignment = align_center
        ws.cell(row=current_row, column=4, value=str(it.get("quality", ""))).alignment = align_center
        ws.cell(row=current_row, column=5, value=str(it.get("chrome_pct", ""))).alignment = align_center
        ws.cell(row=current_row, column=6, value="Yes" if it.get("heat_treatment") else "No").alignment = align_center
        ws.cell(row=current_row, column=7, value=str(it.get("finishing", ""))).alignment = align_left
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_all
        current_row += 1

    ws.cell(row=current_row, column=1, value="HANDLE INFORMATIONS").font = font_bold
    current_row += 1
    write_sub_header(["REFERENCE", "MATERIAL", "PANTONE / PATTERN", "", "", "", ""])
    for it in items_to_render:
        ws.cell(row=current_row, column=1, value=str(it.get("reference", ""))).alignment = align_center
        ws.cell(row=current_row, column=2, value=str(it.get("handle_material", ""))).alignment = align_left
        ws.cell(row=current_row, column=3, value=str(it.get("pantone", ""))).alignment = align_left
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_all
        current_row += 1
    current_row += 1

    # 7. STAMPING & CARTON / EAN
    write_section_header("STAMPING & MASTER CARTON DETAILS")
    ws.cell(row=current_row, column=1, value=f"BLADE OR PRODUCT STAMPING: {data.get('blade_stamping', 'No stamping')}").font = font_bold
    current_row += 1
    write_sub_header(["REFERENCE", "DIM STAMPING (mm)", "CARTON (L×W×H cm)", "EAN 13", "EAN 14 SPCB", "EAN 14 PCB", ""])
    for it in items_to_render:
        dims_carton = f"{it.get('carton_l', '')}×{it.get('carton_w', '')}×{it.get('carton_h', '')}".strip("×")
        ws.cell(row=current_row, column=1, value=str(it.get("reference", ""))).alignment = align_center
        ws.cell(row=current_row, column=2, value=str(it.get("stamping_dim", ""))).alignment = align_center
        ws.cell(row=current_row, column=3, value=dims_carton).alignment = align_center
        ws.cell(row=current_row, column=4, value=str(it.get("ean13", ""))).alignment = align_center
        ws.cell(row=current_row, column=5, value=str(it.get("ean14_inner", ""))).alignment = align_center
        ws.cell(row=current_row, column=6, value=str(it.get("ean14_master", ""))).alignment = align_center
        for col_i in range(1, 8):
            ws.cell(row=current_row, column=col_i).border = border_all
        current_row += 1
    current_row += 1

    # 8. BLOC RÉGLEMENTAIRE AGEC (MOAH / MOSH)
    write_section_header("ARTWORK / AGEC LAW COMPLIANCE")
    agec_text = (
        "In application of the article 112 of the French AGEC Law n°2020-105 of 10 February 2020, "
        "the manufactured products delivered to Tarrerias-Bonjean comply with MOAH / MOSH regulations.\n"
        "Starting Jan 1, 2023: MOAH (1-7 rings) ≤ 1.0% | Starting Jan 1, 2025: MOAH ≤ 0.1%, MOSH ≤ 0.1%"
    )
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=7)
    ws.cell(row=current_row, column=1, value=agec_text).font = font_small
    ws.cell(row=current_row, column=1).alignment = align_left

    # Largeurs de colonnes automatiques.
    #
    # Deux garde-fous : on ignore les cellules fusionnees et on plafonne la
    # largeur. Sans cela, le pave juridique AGEC (300 caracteres dans une
    # cellule fusionnee sur 7 colonnes) donnait une colonne A de 300 de large,
    # et la fiche etait inexploitable a l'impression comme a l'ecran.
    for col in ws.columns:
        longueurs = [
            len(str(cell.value))
            for cell in col
            if cell.value is not None and not isinstance(cell, MergedCell)
            and not any(cell.coordinate in plage for plage in ws.merged_cells.ranges)
        ]
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max(longueurs, default=0) + 3,
                                                         LARGEUR_COLONNE_MIN),
                                                     LARGEUR_COLONNE_MAX)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
